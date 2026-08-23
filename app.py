import streamlit as st
import pandas as pd
import sqlite3
from pathlib import Path
from datetime import datetime, date
from io import BytesIO
import re
import json

# ============================================================
# CONFIGURAÇÃO INICIAL
# ============================================================
st.set_page_config(
    page_title="Dashboard Refugos - WEG UFE",
    layout="wide",
    initial_sidebar_state="expanded",
)

DB_PATH = Path("refugos_weg.db")
TABLE_NAME = "tabela_notas"
PREFS_FILE = Path(".preferencias_usuario.json")

COLUNAS_IMPORTAR = {
    "SEÇÃO": 2, "DEFEITO": 3, "NOTA": 4, "DATA": 5, "TURNO": 6,
    "MATERIAL": 9, "DESCRIÇÃO DO MATERIAL": 10, "CT CAUSADOR": 11,
    "QUANTIDADE": 12, "DESCRIÇÃO DO DEFEITO": 15, "CAUSA": 17,
    "TEXTO DA CAUSA": 18, "CUSTO REFUGO": 20,
}

# ============================================================
# 💾 SALVAR/CARREGAR PREFERÊNCIAS
# ============================================================
def carregar_preferencias():
    if PREFS_FILE.exists():
        try:
            with open(PREFS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return {}

def salvar_preferencias(pref):
    try:
        with open(PREFS_FILE, "w", encoding="utf-8") as f:
            json.dump(pref, f, ensure_ascii=False, indent=2)
    except:
        pass

prefs = carregar_preferencias()

# ============================================================
# 🎨 ESTILO PROFISSIONAL
# ============================================================
st.markdown("""
<style>
.block-container { padding: 1rem 2rem !important; max-width: 100% !important; }
.header-container {
    background: linear-gradient(90deg, #0a192f 0%, #112240 100%);
    padding: 1.2rem 2rem;
    border-radius: 8px;
    color: white;
    margin: 0 0 1.5rem 0;
    box-shadow: 0 2px 8px rgba(0,0,0,0.15);
}
.header-title { font-size: 1.6rem; font-weight: 700; margin: 0; color: #fff; }
.header-subtitle { font-size: 0.95rem; color: #8892b0; margin-top: 0.25rem; }
.metric-card {
    background: #f8fafc;
    border-radius: 8px;
    padding: 1rem 1.25rem;
    border-left: 4px solid;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.metric-card.notas { border-color: #3b82f6; }
.metric-card.qtd  { border-color: #10b981; }
.metric-card.custo { border-color: #f59e0b; }
.metric-card.apq   { border-color: #8b5cf6; }
.metric-value { font-size: 1.8rem; font-weight: 700; line-height: 1.2; }
.metric-label { font-size: 0.85rem; color: #64748b; }
div[data-testid="stDataFrame"] { border-radius: 8px; }
th { background: #f1f5f9 !important; color: #334155 !important; font-weight: 600 !important; }
tr:nth-child(even) { background: #f8fafc; }
section[data-testid="stSidebar"] { background: #f1f5f9; }
</style>

<div class="header-container">
    <h1 class="header-title">⚙️ Dashboard de Refugos — WEG UFE</h1>
    <p class="header-subtitle">Gestão de Apontamentos · Aba "Notas"</p>
</div>
""", unsafe_allow_html=True)

# ============================================================
# 🔧 FUNÇÕES DE LIMPEZA E CONVERSÃO
# ============================================================
def limpar_nota(valor):
    if pd.isna(valor): return ""
    s = str(valor).strip()
    s = re.sub(r'\.0+$', '', s)
    s = re.sub(r'\.\d+$', '', s)
    s = s.replace('.', '').replace(',', '')
    s = re.sub(r'[^0-9]', '', s)
    return s

def limpar_coluna_nota(df, col_nota):
    if col_nota and col_nota in df.columns:
        df[col_nota] = df[col_nota].apply(limpar_nota)
    return df

def formatar_data_br(valor):
    if pd.isna(valor) or str(valor).strip() == "": return ""
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")
    s = str(valor).strip()
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y"]:
        try: return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except: continue
    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(dt): return dt.strftime("%d/%m/%Y")
    except: pass
    return s

def converter_quantidade_inteira(valor):
    if pd.isna(valor) or str(valor).strip() == "": return 0
    s = str(valor).strip().replace(".", "").replace(",", ".")
    try: return int(round(float(s)))
    except: return 0

def converter_custo(valor):
    if pd.isna(valor) or str(valor).strip() == "": return 0.0
    s = str(valor).strip().replace("R$", "").replace(" ", "")
    if "." in s and "," in s: s = s.replace(".", "").replace(",", ".")
    elif "," in s: s = s.replace(",", ".")
    try: return float(s)
    except: return 0.0

# ============================================================
# BANCO DE DADOS
# ============================================================
def get_connection():
    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def table_exists():
    if not DB_PATH.exists(): return False
    conn = get_connection()
    try:
        cur = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (TABLE_NAME,))
        return cur.fetchone() is not None
    finally: conn.close()

def find_column(df, terms):
    for t in terms:
        t = t.lower().strip()
        for c in df.columns:
            if t in str(c).lower().strip(): return c
    return None

def load_data():
    if not table_exists(): return pd.DataFrame()
    conn = get_connection()
    try:
        df = pd.read_sql(f'SELECT rowid, * FROM "{TABLE_NAME}"', conn)
        col_nota = find_column(df, ["nota"])
        df = limpar_coluna_nota(df, col_nota)
        return df
    except Exception as e:
        st.error(f"Erro: {e}")
        return pd.DataFrame()
    finally: conn.close()

# ============================================================
# 🚀 LEITURA OTIMIZADA
# ============================================================
def ler_arquivo_otimizado(arquivo_carregado):
    import openpyxl
    dados_bytes = BytesIO(arquivo_carregado.getvalue())
    wb = openpyxl.load_workbook(dados_bytes, read_only=True, data_only=True)
    
    if "Notas" not in wb.sheetnames:
        raise ValueError(f"Aba 'Notas' não encontrada! Abas: {', '.join(wb.sheetnames)}")
    
    ws = wb["Notas"]
    indices = list(COLUNAS_IMPORTAR.values())
    nomes = list(COLUNAS_IMPORTAR.keys())
    idx_nota = nomes.index("NOTA")
    idx_data = nomes.index("DATA")
    idx_qtd = nomes.index("QUANTIDADE")
    
    dados = []
    for i, linha in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i >= 15000: break
        if len(linha) <= 4 or linha[4] is None or str(linha[4]).strip() == "":
            continue
        
        registro = []
        for j, idx in enumerate(indices):
            if idx < len(linha):
                valor = linha[idx]
                if j == idx_nota: valor = limpar_nota(valor)
                elif j == idx_data: valor = formatar_data_br(valor)
                elif j == idx_qtd: valor = converter_quantidade_inteira(valor)
                elif isinstance(valor, datetime): valor = valor.strftime("%d/%m/%Y")
                registro.append(valor)
            else:
                registro.append(None)
        dados.append(registro)
    
    wb.close()
    df = pd.DataFrame(dados, columns=nomes)
    df = df.drop_duplicates(subset=["NOTA"], keep="first")
    
    for col in ["Observações", "Ação", "Colaborador", "Preparador", "APQ", "TWTP"]:
        df[col] = ""
    df["APQ"] = "Pendente"
    return df

# ============================================================
# 💾 SALVAR DADOS
# ============================================================
def salvar_dados(df_novo):
    col_nota = find_column(df_novo, ["nota"])
    if not col_nota: raise ValueError("Coluna 'Nota' não encontrada!")
    
    df_novo = limpar_coluna_nota(df_novo, col_nota)
    df_novo = df_novo.drop_duplicates(subset=[col_nota], keep="first")
    
    conn = get_connection()
    if table_exists():
        df_antigo = pd.read_sql(f'SELECT * FROM "{TABLE_NAME}"', conn)
        col_nota_ant = find_column(df_antigo, ["nota"])
        
        if col_nota_ant:
            df_antigo = limpar_coluna_nota(df_antigo, col_nota_ant)
            cols_pres = [c for c in 
                ["Observações", "Ação", "Colaborador", "Preparador", "APQ", "TWTP", col_nota_ant]
                if c in df_antigo.columns]
            
            if len(cols_pres) > 1:
                df_pres = df_antigo[cols_pres].copy()
                cols_remover = [c for c in cols_pres if c != col_nota]
                df_novo = df_novo.drop(columns=cols_remover, errors="ignore")
                df_novo = df_novo.merge(df_pres, on=col_nota, how="left")
            
            df_final = pd.concat([df_novo, df_antigo]).drop_duplicates(subset=col_nota, keep="first")
            df_final = limpar_coluna_nota(df_final, col_nota)
            novas = len(set(df_novo[col_nota].unique()) - set(df_antigo[col_nota_ant].unique()))
        else:
            df_final = df_novo
            novas = len(df_novo)
    else:
        df_final = df_novo
        novas = len(df_novo)
    
    df_final.to_sql(TABLE_NAME, conn, if_exists="replace", index=False)
    conn.commit()
    conn.close()
    return len(df_novo), novas

# ============================================================
# 📂 BARRA LATERAL — TODOS OS FILTROS E BOTÃO RESTAURADOS
# ============================================================
with st.sidebar:
    st.header("📂 Importar Dados")
    arq = st.file_uploader("Selecione a Planilha", type=["xlsx", "xlsm", "xls"], label_visibility="collapsed")
    if arq is not None:
        try:
            with st.spinner("Lendo arquivo..."):
                df = ler_arquivo_otimizado(arq)
            if df.empty:
                st.warning("⚠️ Sem dados encontrados.")
            else:
                with st.spinner("Salvando..."):
                    total, novas = salvar_dados(df)
                st.success(f"✅ {total} registros! ({novas} novas notas)")
                st.rerun()
        except Exception as e:
            st.error(f"❌ {str(e)}")

    st.divider()
    st.header("🔍 Filtros")
    
    pesq = st.text_input("Pesquisar Nota", placeholder="Digite o número...")
    
    df_temp = load_data()
    col_data = find_column(df_temp, ["data"])
    tem_data = col_data and not df_temp.empty
    
    f_ano = f_mes = "Todos"
    if tem_data:
        df_temp["__dt__"] = pd.to_datetime(df_temp[col_data], format="%d/%m/%Y", errors="coerce")
        f_ano = st.selectbox("📅 Ano", ["Todos"] + sorted([str(int(x)) for x in df_temp["__dt__"].dt.year.dropna().unique()]))
        f_mes = st.selectbox("📅 Mês", ["Todos"] + sorted([str(int(x)) for x in df_temp["__dt__"].dt.month.dropna().unique()]))
    
    col_secao = find_column(df_temp, ["seção"])
    secoes = ["Todas"] + sorted(df_temp[col_secao].dropna().astype(str).str.strip().unique().tolist()) if col_secao else ["Todas"]
    f_sec = st.selectbox("🏭 Seção", secoes)
    
    col_turno = find_column(df_temp, ["turno"])
    turnos = ["Todos"] + sorted(df_temp[col_turno].dropna().astype(str).str.strip().unique().tolist()) if col_turno else ["Todos"]
    f_turno = st.selectbox("⏰ Turno", turnos)

    st.divider()
    st.header("👁️ Colunas Visíveis")
    
    selecionadas = []
    if not df_temp.empty:
        todas_colunas = [c for c in df_temp.columns if not c.startswith("__")]
        colunas_visiveis_salvas = prefs.get("colunas_visiveis", todas_colunas)
        colunas_visiveis_salvas = [c for c in colunas_visiveis_salvas if c in todas_colunas]
        
        selecionadas = st.multiselect(
            "Escolha quais colunas ver:",
            options=todas_colunas,
            default=colunas_visiveis_salvas,
            key="seletor_colunas"
        )
        
        if selecionadas != colunas_visiveis_salvas:
            prefs["colunas_visiveis"] = selecionadas
            salvar_preferencias(prefs)
            st.toast("✅ Preferência salva!")

    st.divider()
    st.header("⚠️ Administração")
    if st.button("🗑️ Limpar Banco", type="secondary"):
        if DB_PATH.exists():
            DB_PATH.unlink()
            PREFS_FILE.unlink(missing_ok=True)
            st.success("✅ Banco apagado! Reimporte o arquivo.")
            st.rerun()

# ============================================================
# CARREGAR DADOS
# ============================================================
df = load_data()
if df.empty:
    st.info("👋 Bem-vindo! Importe sua planilha pela barra lateral para começar.")
    st.stop()

col_nota = find_column(df, ["nota"])
col_data = find_column(df, ["data"])
col_secao = find_column(df, ["seção", "secao"])
col_turno = find_column(df, ["turno"])
col_qtd = find_column(df, ["quantidade"])
col_custo = find_column(df, ["custo"])
col_apq = find_column(df, ["apq"])
col_obs = find_column(df, ["observação", "observacao", "observações"])
col_acao = find_column(df, ["ação", "acao"])
col_colab = find_column(df, ["colaborador"])

df["__dt__"] = pd.to_datetime(df[col_data], format="%d/%m/%Y", errors="coerce") if col_data else pd.NaT
df["__ano__"] = df["__dt__"].dt.year.astype("Int64")
df["__mes__"] = df["__dt__"].dt.month.astype("Int64")

# ============================================================
# APLICAR FILTROS
# ============================================================
df_f = df.copy()
if pesq and col_nota:
    df_f = df_f[df_f[col_nota].astype(str).str.contains(pesq, case=False, na=False)]
if f_sec != "Todas" and col_secao:
    df_f = df_f[df_f[col_secao].astype(str).str.strip() == f_sec]
if f_turno != "Todos" and col_turno:
    df_f = df_f[df_f[col_turno].astype(str).str.strip() == f_turno]
if f_ano != "Todos":
    df_f = df_f[df_f["__ano__"].astype(str) == f_ano]
if f_mes != "Todos":
    df_f = df_f[df_f["__mes__"].astype(str) == f_mes]

# ============================================================
# 📊 INDICADORES
# ============================================================
st.markdown(f"### 📊 Visão Geral · {len(df_f)} registros")

total_notas = df_f[col_nota].nunique() if col_nota else 0
qtd_total = df_f[col_qtd].apply(converter_quantidade_inteira).sum() if col_qtd else 0
custo_total = df_f[col_custo].apply(converter_custo).sum() if col_custo else 0
total_registros = len(df_f)
concluidas = df_f[col_apq].astype(str).str.lower().isin(["concluída", "concluida", "sim"]).sum() if col_apq else 0
perc_apq = (concluidas / total_registros * 100) if total_registros > 0 else 0

c1, c2, c3, c4 = st.columns(4)

with c1:
    st.markdown(f"""
    <div class="metric-card notas">
        <div class="metric-label">📋 Notas Únicas</div>
        <div class="metric-value">{total_notas:,}</div>
    </div>
    """, unsafe_allow_html=True)

with c2:
    st.markdown(f"""
    <div class="metric-card qtd">
        <div class="metric-label">📦 Quantidade Total</div>
        <div class="metric-value">{int(qtd_total):,}</div>
    </div>
    """, unsafe_allow_html=True)

with c3:
    st.markdown(f"""
    <div class="metric-card custo">
        <div class="metric-label">💰 Custo Total</div>
        <div class="metric-value">R$ {custo_total:,.2f}</div>
    </div>
    """, unsafe_allow_html=True)

with c4:
    st.markdown(f"""
    <div class="metric-card apq">
        <div class="metric-label">✅ APQ Concluídas</div>
        <div class="metric-value">{concluidas} / {total_registros}</div>
        <div style="font-size:0.85rem; color:#6b7280;">{perc_apq:.1f}%</div>
    </div>
    """, unsafe_allow_html=True)

st.divider()

# ============================================================
# 📋 TABELA
# ============================================================
if selecionadas:
    colunas_exibir = selecionadas
else:
    colunas_exibir = [c for c in df_f.columns if not c.startswith("__")]

if "rowid" in df_f.columns and "rowid" not in colunas_exibir:
    colunas_exibir = ["rowid"] + colunas_exibir

df_edit = df_f[colunas_exibir].copy()

cfg = {}
if col_apq and col_apq in colunas_exibir:
    cfg[col_apq] = st.column_config.SelectboxColumn("APQ", options=["Pendente", "Concluída"], required=True)
if col_qtd and col_qtd in colunas_exibir:
    cfg[col_qtd] = st.column_config.NumberColumn("QUANTIDADE", format="%d", min_value=0)
if col_custo and col_custo in colunas_exibir:
    cfg[col_custo] = st.column_config.NumberColumn("CUSTO REFUGO", format="R$ %.2f")

if st.button("💾 Salvar Alterações", type="primary", use_container_width=True):
    try:
        if "rowid" not in df_edit.columns:
            st.warning("⚠️ Reimporte os dados para poder salvar alterações.")
        else:
            conn = get_connection()
            for _, linha in df_edit.iterrows():
                rid = int(linha["rowid"])
                obs = str(linha.get(col_obs, "")).strip() if col_obs else ""
                acao = str(linha.get(col_acao, "")).strip() if col_acao else ""
                colab = str(linha.get(col_colab, "")).strip() if col_colab else ""
                apq = str(linha.get(col_apq, "Pendente")).strip() if col_apq else "Pendente"
                conn.execute(f'UPDATE "{TABLE_NAME}" SET "Observações"=?, "Ação"=?, "Colaborador"=?, "APQ"=? WHERE rowid=?',
                    (obs, acao, colab, apq, rid))
            conn.commit()
            conn.close()
            st.success("✅ Alterações salvas com sucesso!")
            st.rerun()
    except Exception as e:
        st.error(f"❌ Erro ao salvar: {e}")

st.data_editor(
    df_edit, use_container_width=True, hide_index=True, num_rows="fixed",
    column_config=cfg, key="tabela", height=500
)
